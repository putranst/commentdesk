from openpyxl import load_workbook

path = '/Users/putra/Documents/PROJECTS/DM/duniameutya_last_10_posts_comments.xlsx'
wb = load_workbook(path)

# The public profile view exposed these account-owned posts from Aug 1 onward,
# plus the three posts already verified earlier in this conversation.
keep = {
    'Bahlil_Book': ('2026-08-07', 'https://www.instagram.com/p/DbvXVj9GqVD/'),
    'Bahlil_Animation': ('2026-08-06', 'https://www.instagram.com/p/DbudwrNkbTV/'),
    'RANA_Anak': ('2026-08-05', 'https://www.instagram.com/p/DbpzG0ykVoJ/'),
    'KIP_Informasi': ('2026-08-03', 'https://www.instagram.com/p/DbmUBwREwkB/'),
    'Riset_BRIN': ('2026-08-01+', 'https://www.instagram.com/p/DbtFV4_hKjO/'),
    'Hoaks_Provokasi': ('2026-08-01+', 'https://www.instagram.com/p/DbqDgSEwj-4/'),
    'Bahlil_Greeting': ('2026-08-01+', 'https://www.instagram.com/p/Dbuqj8MxoH0/'),
}
for sheet in list(wb.sheetnames):
    if sheet not in keep:
        del wb[sheet]
for sheet, (date, url) in keep.items():
    ws = wb[sheet]
    ws['G1'] = 'Post date'
    ws['H1'] = date
    ws['G2'] = 'Source URL'
    ws['H2'] = url
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 48
    ws['G1'].font = ws['G2'].font = ws['A1'].font.copy(bold=True)
wb.save(path)
check = load_workbook(path, read_only=True)
assert check.sheetnames == list(keep)
assert all(ws.max_row == 26 for ws in check.worksheets)
print('verified', len(check.sheetnames), 'tabs x 25 comments =', len(check.sheetnames)*25)
print('sheets:', ', '.join(check.sheetnames))
print('path:', path)
