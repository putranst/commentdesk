from openpyxl import load_workbook

path = '/Users/putra/Documents/PROJECTS/DM/duniameutya_last_10_posts_comments.xlsx'
wb = load_workbook(path)

extra = {
'Bahlil_Book': [
'ulang tahun sekaligus launching buku... momentumnya pas bgt pak 🎂 semoga sukses selalu',
'perjalanan 50 tahun pasti ga gampang... salut, semoga ceritanya menginspirasi banyak orang',
'ditunggu karya berikutnya pak... semoga selalu diberi kesehatan dan kelancaran',
'buku tentang perjalanan hidup pasti banyak pelajarannya yaa... semoga laris dan bermanfaat',
'keren pak, dari pengalaman jadi cerita, dari cerita jadi inspirasi buat generasi muda',
'barakallah fi umrik Pak Bahlil... semoga makin amanah dan dekat dengan rakyat 🤲',
'50 tahun bukan akhir perjalanan... justru semoga makin banyak gebrakan buat negeri',
'judul bukunya menarik bgt... 10 karya nyata, semoga benar2 jadi catatan yg bermanfaat',
'selamat ulang tahun pak... sehat terus, energinya jangan habis buat Indonesia 🔥',
'perjalanan dari timur sampai ke panggung nasional... luar biasa pak, tetap rendah hati',
'kayaknya bukunya wajib dibaca anak muda yg lagi mulai berjuang dari bawah',
'semoga buku ini membuka banyak cerita baik dan jadi warisan inspirasi untuk generasi berikutnya',
'ulang tahun ke 50, buku baru, semangat baru... sehat dan sukses terus pak 🙏',
'karya nyata memang lebih kuat daripada sekadar kata2... ditunggu karya selanjutnya pak',
'Happy birthday Pak Bahlil... semoga setiap langkah pengabdian selalu diberi keberkahan'
],
'Bahlil_Animation': [
'ini siapa yg bikin animasinya... kreatif bgt 😄 selamat ulang tahun Pak Bahlil',
'baru kali ini ucapan ultah menteri sekreatif ini wkwk... keren bu',
'versi kartun aja auranya udah pemimpin bgt 🔥 sehat terus pak',
'pak ketum di animasi keliatan muda terus... rahasianya apa nih 😆',
'ucapan kreatif, doa terbaik... selamat ulang tahun Pak Bahlil 🙏',
'energi terang untuk negeri... semoga realisasinya juga makin terang ya pak',
'bagus nih kontennya, ringan tapi tetep ada pesan dan doanya',
'yg di sebelah Pak Ketum bikin penasaran... harus swipe sampe akhir ternyata 😄',
'sehat2 selalu pak, semoga makin banyak kebijakan yg bermanfaat buat rakyat',
'kalo semua ucapan ultah sekreatif ini, notif pasti makin ditunggu wkwk',
'ulang tahun dengan gaya beda... happy birthday pak, sukses terus',
'animasi lucu, pesannya serius... sukaa 👍',
'semoga tetap jadi pemimpin yg menginspirasi dan dekat dengan masyarakat',
'keren bu, ucapan ulang tahunnya niat banget... sehat terus Pak Bahlil',
'pak menteri paling gemes hari ini 😄 semoga sehat, berkah dan amanah'
],
'RANA_Anak': [
'ruang digital aman dimulai dari rumah... orang tua jg harus belajar bareng anak',
'konten anak sekarang banyak banget, jadi pilih2 dan dampingi itu wajib yaa',
'semoga ada panduan praktis buat ibu2, bukan cuma jargon yg susah dipahami',
'gerakan yg sangat dibutuhkan di zaman sekarang... anak2 harus dilindungi',
'kadang anak lebih ngerti teknologi daripada orang tuanya 😅 yuk belajar sama2',
'jangan tunggu anak jadi korban baru bergerak... pencegahan itu penting bgt',
'semoga sekolah2 ikut aktif ngajarin etika dan keamanan digital',
'buat para orang tua, ini pengingat penting supaya ga cuek sama aktivitas online anak',
'bagus bgt... dunia maya juga harus jadi tempat yg aman untuk tumbuh dan belajar',
'perlindungan anak bukan cuma urusan pemerintah, keluarga dan platform jg punya peran',
'buibu pasti setuju... anak dikasih hp memang butuh aturan dan pendampingan',
'semoga programnya sampai ke daerah2 dan bukan hanya kota besar',
'anak2 perlu ruang berekspresi, tapi tetap ada pagar pengamannya',
'keren langkahnya... semoga konsisten dan hasilnya terasa di keluarga',
'konten positif buat anak harus diperbanyak, jangan yg bikin mereka takut atau bingung'
],
'KIP_Informasi': [
'kalau informasi publik mudah dicari, masyarakat juga ga gampang percaya kabar liar',
'akses informasi yg jelas bisa mencegah salah paham... ini penting bgt',
'semoga semua lembaga makin terbuka dan responsif menjawab pertanyaan warga',
'di era AI, verifikasi informasi memang harus jadi kebiasaan bersama',
'info resmi harus hadir sebelum rumor berkembang kemana2 🙏',
'keterbukaan bukan cuma upload dokumen, tapi juga bahasa yg mudah dipahami',
'semoga target 75 bisa tercapai dan kualitas pelayanannya ikut naik',
'masyarakat perlu tau haknya untuk mendapatkan informasi yg benar',
'bagus... kepercayaan publik dibangun dari informasi yg akurat dan konsisten',
'jangan sampai warga harus punya kenalan dulu baru dapat informasi yaa',
'deepfake bikin kita makin sadar pentingnya sumber resmi dan cek fakta',
'semoga sengketa informasi bisa diselesaikan cepat dan adil',
'informasi publik itu hak, bukan bonus... semoga terus diperkuat',
'layanan digital yg simpel pasti sangat membantu warga yg jauh dari pusat',
'langkah baik untuk menjaga demokrasi tetap sehat di tengah banjir informasi'
],
'Riset_BRIN': [
'karya anak bangsa harus dikasih ruang dan dukungan seperti ini... bangga 🇮🇩',
'semoga perisetnya juga dapat fasilitas dan pendanaan yg berkelanjutan',
'hasil riset harus nyambung dengan kebutuhan masyarakat sehari2',
'keren... dari laboratorium menuju industri, semoga jalannya makin lancar',
'jangan lupa libatkan kampus dan komunitas lokal juga yaa',
'kalau riset berkembang, lapangan kerja dan solusi baru juga ikut tumbuh',
'buibu pasti senang kalau inovasi riset bisa bantu kesehatan dan pendidikan anak',
'Indonesia punya banyak talenta, tinggal ekosistemnya yg harus terus diperkuat',
'semoga anak2 muda melihat bahwa jadi peneliti itu keren dan punya masa depan',
'riset strategis memang harus jadi prioritas negara... lanjutkan terus',
'hasil penelitian yg masuk industri bisa bikin kita lebih mandiri',
'bagus banget kalau periset langsung bisa berdialog dengan pembuat kebijakan',
'jangan berhenti di pameran yaa... ditunggu produk dan solusi nyatanya',
'negara maju dibangun dari ilmu, bukan cuma slogan... semangat para periset',
'semoga dukungan untuk sains konsisten lintas tahun dan lintas pemerintahan'
],
'Hoaks_Provokasi': [
'kadang yg bikin rusak bukan beritanya aja, tapi emosi yg muncul setelah baca',
'cek dulu sebelum percaya... jempol kita bisa berdampak panjang lho',
'berita provokasi biasanya bikin kita buru2 pilih pihak, padahal faktanya belum jelas',
'jangan sampai keluarga dan teman jadi musuhan gara2 postingan yg belum dicek',
'kalo judulnya terlalu heboh, biasanya justru perlu dicurigai 😅',
'buibu, kalo ada berita anak hilang atau kesehatan di WA, cek sumber dulu yaa',
'beda pilihan tetap saudara... jangan kasih hoaks ruang untuk memecah kita',
'platform digital ramai, tapi akal sehat tetap harus nomor satu',
'semoga makin banyak konten edukasi soal cek fakta yg gampang dipraktikkan',
'hoaks bisa menyebar cepat, klarifikasi jg harus cepat dan mudah ditemukan',
'jangan ikut menyebarkan sesuatu cuma karena takut ketinggalan info',
'kalau belum yakin, simpan dulu... ga harus langsung share kok',
'ruang digital tenang dimulai dari kita yg mau membaca lebih lengkap',
'provokasi sering dibungkus seolah2 fakta... makanya sumber itu penting',
'jaga persatuan, jaga percakapan, jaga Indonesia 🙏'
],
'Bahlil_Greeting': [
'barakallah fi umrik Pak Bahlil... semoga selalu sehat dan amanah 🤲',
'ucapan terus mengalir... semoga doa baiknya kembali ke semua yg mendoakan',
'semoga di usia 50 makin matang, makin bijak dan makin bermanfaat',
'selamat ulang tahun pak... tetap semangat melangkah dari Timur',
'notif penuh tapi semoga hatinya juga penuh rasa syukur ya pak 😄',
'panjang umur dan sehat selalu untuk Pak Menteri 🎂',
'semoga keluarga juga selalu diberi kesehatan dan kebahagiaan',
'kalo kue udah dipotong jangan lupa kirim virtual ya pak wkwk',
'ulang tahun itu pengingat perjalanan... semoga langkah ke depan makin berkah',
'sukses terus pak, semoga energi positifnya menular ke generasi muda',
'50 tahun, semoga makin banyak pengalaman yg jadi manfaat bagi negeri',
'selamat hari lahir pak... semoga semua urusan dimudahkan',
'yg penting sehat dulu pak, urusan gebrakan menyusul 😄',
'ucapan sederhana tapi doanya panjang... sehat dan bahagia selalu pak',
'Happy birthday Pak Bahlil, semoga selalu dekat dengan masyarakat dan dicintai rakyat'
],
'Anti_Hoaks': [
'jangan langsung percaya screenshot tanpa konteks yaa... sering banget menyesatkan',
'kalo sumbernya akun anonim, tahan dulu jempolnya 😅',
'cek tanggal berita juga penting, kadang berita lama diangkat lagi',
'orang tua perlu dibekali cara cek fakta yg sederhana dan praktis',
'konten edukasi begini harus sering muncul di beranda kita',
'ga semua yg viral itu benar, ga semua yg benar itu langsung viral',
'berhenti sebentar sebelum share bisa menyelamatkan banyak orang dari salah info',
'judul provokatif memang menarik klik, tapi belum tentu isinya valid',
'semoga anak2 juga terbiasa bertanya: sumbernya dari mana?',
'kalau beda pendapat, bahas datanya... bukan serang orangnya',
'literasi digital bukan cuma buat anak muda, orang dewasa jg perlu',
'grup keluarga aman kalo semua mau cek dulu sebelum forward 😄',
'berita kesehatan, hukum dan keuangan jangan asal percaya yaa',
'kita semua punya peran menjaga timeline tetap waras',
'konten yg informatif dan menenangkan jauh lebih bermanfaat daripada provokasi'
],
'Perlindungan_Publik': [
'publik berhak merasa aman saat menggunakan teknologi dan media sosial',
'jangan normalisasi menyebarkan wajah atau rekaman orang tanpa izin',
'semoga aturan perlindungan datanya makin kuat dan penindakannya jelas',
'ibu2 pasti khawatir soal foto anak yg tersebar... ini penting bgt',
'platform harus ikut bertanggung jawab, bukan cuma pengguna yg disalahkan',
'kalau ada pelanggaran, proses lapornya harus gampang dan cepat',
'teknologi seharusnya memudahkan hidup, bukan bikin orang kehilangan rasa aman',
'bagus kalau edukasi privasi dimulai sejak sekolah',
'konten viral bukan berarti bebas dipakai seenaknya yaa',
'semoga masyarakat makin paham soal hak digitalnya sendiri',
'perlindungan publik harus berlaku untuk semua, termasuk kelompok rentan',
'aturan yg jelas bikin inovasi juga lebih dipercaya masyarakat',
'jangan tunggu kasus besar baru serius melindungi privasi warga',
'ruang digital aman perlu kolaborasi pemerintah, platform, keluarga dan sekolah',
'langkah baik... semoga terus ada pengawasan dan evaluasi terbuka'
],
'Kata_Mereka': [
'kalau warga diajak ngobrol, kebijakan biasanya jadi lebih nyambung dengan kebutuhan',
'semoga suara dari daerah juga sering ditampilkan, bukan hanya dari kota besar',
'dengerin masyarakat itu penting, tapi follow up-nya juga jangan hilang yaa',
'kadang solusi terbaik datang dari pengalaman warga sehari2',
'bagus bgt ada ruang dialog seperti ini... semoga rutin dilakukan',
'masyarakat perlu tahu bahwa masukannya benar2 dicatat dan diproses',
'buibu juga ingin didengar soal keamanan digital dan pendidikan anak',
'komunikasi dua arah bikin pemerintah lebih dekat dengan rakyat',
'jangan takut menerima kritik, karena dari situ biasanya perbaikan dimulai',
'konten seperti ini bikin warga merasa lebih dilibatkan 🙏',
'semoga hasil diskusinya bukan hanya jadi dokumentasi, tapi ada kabar lanjutannya',
'beda daerah beda masalah... semoga kebijakannya bisa lebih fleksibel',
'ruang ngobrol yg terbuka bisa mengurangi salah paham antara pemerintah dan masyarakat',
'keren... dengar dulu, pahami, baru susun solusi',
'partisipasi masyarakat memang harus jadi bagian dari kerja pemerintahan'
]
}

for name, comments in extra.items():
    ws = wb[name]
    start = ws.max_row + 1
    for i, comment in enumerate(comments, start):
        ws.append([i, comment, 'Belum digunakan', '', 'Tambahan batch'])
    ws.auto_filter.ref = f'A1:E{ws.max_row}'

wb.save(path)
print('updated', path)
for ws in wb.worksheets:
    print(ws.title, ws.max_row - 1)
# verify reopen
check = load_workbook(path, read_only=True)
assert all(ws.max_row == 26 for ws in check.worksheets)
print('verified: 10 sheets x 25 comments')
