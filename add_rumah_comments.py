from openpyxl import load_workbook
from openpyxl.styles import Font

path='/Users/putra/Documents/PROJECTS/DM/duniameutya_last_10_posts_comments.xlsx'
wb=load_workbook(path)
name='Rumah_Rakyat'
if name in wb.sheetnames:
    del wb[name]
ws=wb.create_sheet(name, 0)
headers=['No.','Comment','Status','Used by','Notes']
ws.append(headers)
comments=[
'rumah layak + internet aman... kebutuhan zaman sekarang bgt 🙏',
'bagus bgt, rumah rakyat jgn cuma soal bangunan tapi koneksi jg penting',
'buibu pasti setuju... anak belajar online butuh internet yg stabil yaa',
'semoga akses internetnya bener2 sampai ke rumah2 rakyat, bukan cuma di kota besar',
'keren... rumah nyaman, internet lancar, keluarga bisa belajar dan usaha dari rumah',
'program tiga juta rumah semoga makin kuat dengan dukungan infrastruktur digital',
'jaman sekarang rumah tanpa internet rasanya belum lengkap 😅',
'semoga harga internetnya jg terjangkau ya bu, jangan koneksi ada tapi mahal',
'rumah layak itu hak semua orang... konektivitas jg sudah jadi kebutuhan dasar',
'bagus langkahnya, digitalisasi harus ikut dibangun dari kawasan hunian baru',
'buat anak2 yg tinggal jauh dari pusat kota, internet bisa buka akses pendidikan luas bgt',
'semoga pelaksanaannya tepat sasaran dan kualitas jaringannya bener2 bagus',
'rumah rakyat + internet cepat = peluang kerja dan usaha makin terbuka 🔥',
'jangan lupa edukasi keamanan digital buat keluarga yg baru dapat akses internet yaa',
'konektivitas penting untuk sekolah, layanan kesehatan, administrasi dan usaha warga',
'keren bu... pembangunan fisik dan digital memang harus jalan bareng',
'semoga semua keluarga bisa menikmati internet yg aman, stabil dan ga putus2',
'kalau jaringan bagus, UMKM rumahan juga bisa naik kelas 🙏',
'program yg menyentuh kebutuhan sehari2 masyarakat... semoga konsisten dikawal',
'rumahnya dibangun, akses digitalnya jg disiapkan dari awal... mantap',
'buat ibu2, internet aman bisa bantu cari informasi pendidikan dan kesehatan anak',
'semoga daerah terpencil jg dapat perhatian yg sama, jangan cuma wilayah perkotaan',
'pembangunan masa kini memang harus mikirin koneksi... bukan hanya tembok dan atap',
'bagus banget... semoga manfaatnya benar2 terasa sampai keluarga paling bawah',
'rumah nyaman, koneksi aman, masa depan anak lebih terbuka 🇮🇩'
]
for i,c in enumerate(comments,1): ws.append([i,c,'Belum digunakan','', 'Post baru'])
ws['G1']='Post date'; ws['H1']='2026-08-01+'; ws['G2']='Source URL'; ws['H2']='https://www.instagram.com/p/DbemzciRYbI/'
ws['G1'].font=ws['G2'].font=Font(bold=True)
for cell in ws[1]: cell.font=Font(bold=True)
ws.freeze_panes='A2'; ws.auto_filter.ref='A1:E26'
ws.column_dimensions['A'].width=8; ws.column_dimensions['B'].width=85; ws.column_dimensions['C'].width=18; ws.column_dimensions['D'].width=20; ws.column_dimensions['E'].width=18; ws.column_dimensions['G'].width=15; ws.column_dimensions['H'].width=48
for row in ws.iter_rows(): row[1].alignment=__import__('openpyxl').styles.Alignment(wrap_text=True, vertical='top')
wb.save(path)
check=load_workbook(path,read_only=True)
assert len(check.sheetnames)==8 and all(s.max_row==26 for s in check.worksheets)
print('verified',len(check.sheetnames),'tabs x 25 comments =',len(check.sheetnames)*25)
print(check.sheetnames)
print(path)

# Copy a second time with a clear filename in the same DM folder.
import shutil
out='/Users/putra/Documents/PROJECTS/DM/duniameutya_from_Dbem_to_latest_25_comments.xlsx'
shutil.copy2(path,out)
print(out)
